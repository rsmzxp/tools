#encoding=utf-8
import os
import MySQLdb
import openpyxl
import argparse
from multiprocessing import Pool
from functools import partial

def my_thread(in_file,ref,n):
    result = {}
    files=os.listdir(in_file)
    parse = list(zip([in_file]*len(files),[ref]*len(files),files))
    pool = Pool(int(n))
    values = pool.starmap(deal_data,parse)
    pool.close()
    pool.join()
    for value in values:
        result[value[0]] = value[1:]
    return result

def deal_data(path,ref,file):
    #从数据库中获取结果，并进行判断，突变型是否一致，返回一个字典，样本名为key，单倍型信息为value
    print(file)
    result = [file]
    conn = MySQLdb.connect("localhost","dbuser", "dbuser2018","public",charset="utf8")
    cursor = conn.cursor()
    with open(path+'/'+file,'r') as f:
        lines=f.readlines()
        my_key = file.replace('.vcf', '')
        # print(my_key)
        types=[]
        for line in lines:
            if line.startswith('chrY'):
                type=line.strip('\n').split('\t')
                infos=type[9].split(':')
                if infos[0] == '1/1' or infos[0]=='1|1':
                    if int(infos[2])>=1:
                        # print(type[0]+':'+type[1]+'->'+type[2])
                        singletype,mutants,anc=get_info_from_databse(cursor,ref,type[1])
                        if type[4].find(',') == -1:
                            if len(singletype)!=0 and type[4] in mutants:
                                # print(type[0] + ':' + type[1] + '->' + type[2])
                                types+=singletype
                                print(type[1])
                        else:
                            if len(singletype)!=0 and type[4].split(',')[0] in mutants:
                                # print(type[0] + ':' + type[1] + '->' + type[2])
                                types+=singletype
                                # print(type[1])
        result+=sorted(list(set(types)))
    cursor.close()
    conn.close()
    return result

def save_xls(result,out_file):
    print('saving')
    wb=openpyxl.Workbook()
    ws=wb.active
    # ws1=wb.create_sheet('Mysheet')
    for x,name in enumerate(result.keys()):
        ws.cell(column=2*x+1,row=1,value=name)
        for y,value in enumerate(result[name]):
            ws.cell(column=2*x+1,row=y+2,value=value)
    wb.save(out_file+'result.xlsx')

def annova_vcf(file,ref):
    #输入一个vcf文件,根据位置，对其进行单倍型注释
    out_data=[]
    conn = MySQLdb.connect("localhost", "dbuser", "dbuser2018", "public", charset="utf8")
    cursor=conn.cursor()
    with open(file,'r') as f:
        lines = f.readlines()
        for line in lines:
            # count+=1
            # print(count)
            if line.startswith('##'):
                out_data.append(line)
            else:
                temp = line.strip('\n').split('\t')
                if line.startswith('chrY'):
                    type,mutant,anc=get_info_from_databse(cursor,ref,temp[1])
                    if  len(mutant)==0:
                        new_temp=temp[0:5]+['.','.','.']+temp[5:]
                    elif len(mutant)==1:
                        if mutant[0] in temp[4]:
                            new_temp = temp[0:5] + [anc[0],mutant[0], ','.join(type)] + temp[5:]
                        else:
                            new_temp = temp[0:5] + [anc[0],mutant[0], '.'] + temp[5:]
                    elif len(mutant)==2:
                        if mutant[0] in temp[4]:
                            new_temp = temp[0:5] + [anc[0],mutant[0], type[0]] + temp[5:]
                        elif mutant[1] in temp[4]:
                            new_temp = temp[0:5] + [anc[0],mutant[1], type[1]] + temp[5:]
                        else:
                            new_temp = temp[0:5] + [','.join(anc),','.join(mutant), '.'] + temp[5:]
                    else:
                        new_temp=temp[0:5]+['.','.','.']+temp[5:]
                else:
                    new_temp = temp[0:5] + ['ANC','DER', 'subgroup name'] + temp[5:]
                out_data.append('\t'.join(new_temp)+'\n')
    cursor.close()
    conn.close()
    return out_data

def get_type_from_databse(cursor,ref,pos):
    type = []
    mutant = []
    sql1 = 'select * from name_type_0628 where gene_pos_' + ref + '=%s;'
    cursor.execute(sql1, [pos])
    infos1 = cursor.fetchall()
    for i in infos1:
        type.append(i[2])
        if len(i[6].split('->')[1]) < 2:
            mutant.append(i[6].split('->')[1])
    return list(set(type)),list(set(mutant))

def get_info_from_databse(cursor,ref,pos):
    #传入的参数pos,基因的位置信息
    #返回值type单倍体型，mutant突变型
    type=[]
    mutant=[]
    anc=[]
    sql1='select * from name_type_0628 where gene_pos_'+ref+'=%s;'
    sql2='select * from yfull where Build'+ref+'=%s'
    cursor.execute(sql1,[pos])
    infos1 = cursor.fetchall()
    # print(infos)
    for i in infos1:
        type.append(i[2])
        if len(i[6].split('->')[1])<2:
            mutant.append(i[6].split('->')[1])
            anc.append(i[6].split('->')[0])
    cursor.execute(sql2, [pos])
    infos2 = cursor.fetchall()
    # print(infos)
    for i in infos2:
        type.append(i[2])
        mutant.append(i[6])
        anc.append(i[5])
    type=list(set(type))
    mutant=list(set(mutant))
    # print(key,mutant)
    return type,mutant,anc

if __name__=='__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-m', '--model', required=True,  help='1.提取样本的单倍型信息 2.注释vcf文件')
    parser.add_argument('-i', '--in_vcf', required=True, help='输入的vcf文件')
    parser.add_argument('-o', '--out_vcf', required=True, help='输出的结果文件')
    parser.add_argument('-r', '--reference', required=True, help='37 or 38')
    parser.add_argument('-t', '--thread', default=1, help='进程数')
    args = parser.parse_args()
    # save_json('result.json',deal_data())
    # save_xls()
    if args.model=='1':
        if args.in_vcf and args.out_vcf and args.reference:
            in_file = args.in_vcf
            out_file = args.out_vcf
            ref = args.reference
            n = args.thread
            print('strating')
            result=my_thread(in_file,ref,n)
            save_xls(result,out_file)
    elif args.model=='2':
        if args.in_vcf and args.out_vcf and args.reference:
            in_file = args.in_vcf
            out_file = args.out_vcf
            ref = args.reference
            result = annova_vcf(in_file,ref)
            print('strating')
            with open(out_file+'.vcf', 'w') as f:
                f.writelines(result)
    else:
        print('缺少参数')
