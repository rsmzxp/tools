import pymysql
import json
import openpyxl
import csv

def save_xls(data):
    #传入的参数是一个json文件，字典，但是excel太慢了
    wb=openpyxl.Workbook()
    ws=wb.active
    # ws1=wb.create_sheet('Mysheet')
    result=data
    for x,name in enumerate(result.keys()):
        ws.cell(column=2*x+1,row=1,value=name)
        for y,value in enumerate(result[name]):
            ws.cell(column=2*x+1,row=y+2,value=value)
    wb.save('result.xlsx')

def save_csv(data):
    #传入的参数是一个二维数组
    with open('result.csv','a',newline='') as f:
    csv_write = csv.writer(f, dialect='excel')
    for i in data:
        csv_write.writerow(i)
    print('写入成功')
        

def save_json(name,data):
    with open(name,'w') as f:
        json.dump(data,f)

def read_json(name):
    with open(name,'r') as f:
        result=json.load(f)
    return result
